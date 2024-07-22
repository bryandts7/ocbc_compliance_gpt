import os

import docx2txt
import fitz
from langchain_core.documents import Document
from openpyxl import load_workbook
from paddleocr import PaddleOCR
from utils.constants import OCR_CONFIDENCE_THRESHOLD
from utils.utils import ensure_metadata_fields


def extract_text_and_images_from_page(doc, page, ocr):
    text = page.get_text()
    image_text = ""
    image_list = page.get_images(full=True)
    
    # Iterate through all images found on the page
    for image_info in image_list:
        xref = image_info[0]
        image_dict = doc.extract_image(xref)
        image_bytes = image_dict['image']

        # Use PaddleOCR to extract text from the image
        ocr_result = ocr.ocr(image_bytes)
        
        # Check if OCR result is valid before processing
        if ocr_result and ocr_result != [None]:
            for result in ocr_result:
                for res in result:
                    text_tuple = res[1]
                    text_string = text_tuple[0]
                    text_confidence = text_tuple[1]  # For confidence threshold

                    if text_confidence > OCR_CONFIDENCE_THRESHOLD:
                        image_text += text_string + '\n'
    
    # Combine page text and image text
    return text + "\n" + image_text

def extract_text_from_pdf(ocr, pdf_path):
    try:
        doc = fitz.open(pdf_path)
        combined_text = ""
        page_data_list = []
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            page_text = extract_text_and_images_from_page(doc, page, ocr)
            combined_text += page_text + "\n"
            page_data_list.append({'page_number': page_num + 1, 'text': page_text})
        return combined_text, page_data_list
    except:
        return "", []

def extract_text_from_docx(docx_path):
    return docx2txt.process(docx_path)


def extract_text_from_excel(excel_path):
    wb = load_workbook(excel_path)
    text_data = []

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    text_data.append(str(cell))
    
    return ' '.join(text_data)


def extract_text(ocr, file_path):
    extension = file_path.lower()
    if extension.endswith('.pdf'):
        return extract_text_from_pdf(ocr, file_path)
    elif extension.endswith('.docx'):
        text = extract_text_from_docx(file_path)
        return text, []
    elif extension.endswith('.xlsx'):
        text = extract_text_from_excel(file_path)
        return text, []
    else:
        print("file not supported", file_path)
        return "", []
        # raise ValueError("File not supported!")


def add_contextual_header(document):
    header = f"DOC NAME: {document.metadata['title']} | {document.metadata['type_of_regulation']}\n\n"
    document.page_content = header + document.page_content
    return document


def process_documents(ocr, data_dir, metadata):
    text_data = []
    required_fields = [
        "file_id", "title", "file_name", "file_link", "date", "type_of_regulation", "sector",
        "standardized_extracted_file_name", "standardized_file_name"
    ]

    for subdir, _, files in os.walk(data_dir):
        for file in files:
            file_path = os.path.join(subdir, file)
            text, page_metadata = extract_text(ocr, file_path)
            
            # Extract the ID from the filename (assuming it's the last 8 characters before the extension)
            filename, _ = os.path.splitext(file)
            file_id = filename[-8:]

            file_metadata = metadata.get(file_id, {})
            file_metadata = ensure_metadata_fields(file_metadata, required_fields)
            
            if page_metadata:
                for page in page_metadata:
                    page_specific_metadata = file_metadata.copy()
                    page_specific_metadata['page_number'] = page['page_number']
                    document = Document(page_content=page['text'], metadata=page_specific_metadata)
                    document = add_contextual_header(document)
                    text_data.append(document)
            else:
                document = Document(page_content=text, metadata=file_metadata)
                document = add_contextual_header(document)
                text_data.append(document)

    return text_data


if __name__ == "__main__":
  ocr = PaddleOCR(use_angle_cls=True, lang='en')
  pdf = 'data_pool/pbi/pbi-2_15_pbi_2000-20032004-bank_indonesia_regulation_no2_15_pbi_2000_dated_june_12_2000___bank_indonesia_regulation_no2_16_pbi_2000___concerning_amendment_to_decree_of_the_board_of_managing_directors_of_bank_indonesia_no31_150_kep_dir_dated_uxl.pdf'
  extracted_text = extract_text_from_pdf(ocr, pdf)
  print("result", extracted_text)